"""
Excel Report Generator for Relationship Engine
Produces weekly/daily .xlsx reports: pipeline summary, signal digest,
outreach tracker, top opportunities.
"""

import os
import sys
import sqlite3
from datetime import datetime, timedelta

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config  # noqa: F401

from graph_engine import get_db_path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# =============================================================================
# STYLE CONSTANTS
# =============================================================================

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_SUBHEADER = Font(name="Arial", bold=True, size=10)
FONT_BODY = Font(name="Arial", size=10)
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_DATE = Font(name="Arial", size=9, italic=True, color="666666")
FONT_BLUE = Font(name="Arial", size=10, color="0000FF")

FILL_HEADER = PatternFill("solid", fgColor="2F5496")
FILL_ALT_ROW = PatternFill("solid", fgColor="D6E4F0")
FILL_HIGH = PatternFill("solid", fgColor="FFC7CE")
FILL_MED = PatternFill("solid", fgColor="FFEB9C")
FILL_LOW = PatternFill("solid", fgColor="C6EFCE")
FILL_SECTION = PatternFill("solid", fgColor="E2EFDA")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def _style_header_row(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def _style_data_row(ws, row, col_count, alt=False):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_BODY
        cell.border = THIN_BORDER
        if alt:
            cell.fill = FILL_ALT_ROW


def _auto_width(ws, min_width=10, max_width=40):
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


# =============================================================================
# DATA QUERIES
# =============================================================================

def _get_conn(db_path=None):
    if db_path is None:
        db_path = get_db_path()
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def _fetch_pipeline(conn):
    cur = conn.cursor()
    cur.execute("""
        SELECT
            c.name, c.status, c.industry,
            COALESCE(c.opportunity_score, 0) as opp_score,
            MAX(o.outreach_date) as last_outreach,
            MAX(o.outcome) as last_outcome,
            COUNT(o.id) as total_touches,
            d.deal_type, d.status as deal_status
        FROM companies c
        LEFT JOIN outreach_log o ON c.id = o.target_company_id
        LEFT JOIN (
            SELECT company_id, deal_type, status,
                   ROW_NUMBER() OVER (PARTITION BY company_id ORDER BY COALESCE(started_date, created_at) DESC) as rn
            FROM deals WHERE status NOT IN ('lost', 'dead')
        ) d ON c.id = d.company_id AND d.rn = 1
        WHERE c.status IN ('high_growth_target', 'prospect', 'active_client', 'watching')
        GROUP BY c.id
        ORDER BY COALESCE(c.opportunity_score, 0) DESC
    """)
    return [dict(r) for r in cur.fetchall()]


def _fetch_signals(conn, days_back=30):
    cur = conn.cursor()
    signals = []

    cur.execute("""
        SELECT f.event_date as date, 'Funding' as type, c.name,
               COALESCE(f.round_type, 'Unknown') || ' - $' || COALESCE(CAST(CAST(f.amount AS INTEGER) AS TEXT), '?') as detail,
               f.lead_investor as extra, c.status
        FROM funding_events f
        JOIN companies c ON f.company_id = c.id
        WHERE f.event_date >= date('now', ? || ' days')
        ORDER BY f.event_date DESC
    """, (f"-{days_back}",))
    signals.extend([dict(r) for r in cur.fetchall()])

    cur.execute("""
        SELECT h.signal_date as date, 'Hiring' as type, c.name,
               REPLACE(h.signal_type, '_', ' ') as detail,
               h.relevance as extra, c.status
        FROM hiring_signals h
        JOIN companies c ON h.company_id = c.id
        WHERE h.signal_date >= date('now', ? || ' days')
        ORDER BY h.signal_date DESC
    """, (f"-{days_back}",))
    signals.extend([dict(r) for r in cur.fetchall()])

    cur.execute("""
        SELECT l.lease_expiry as date, 'Lease Expiry' as type, c.name,
               COALESCE(CAST(CAST(l.square_feet AS INTEGER) AS TEXT), '?') || ' SF' as detail,
               COALESCE(b.address, '') as extra, c.status
        FROM leases l
        JOIN companies c ON l.company_id = c.id
        LEFT JOIN buildings b ON l.building_id = b.id
        WHERE l.lease_expiry BETWEEN date('now') AND date('now', '+12 months')
        ORDER BY l.lease_expiry ASC
    """)
    signals.extend([dict(r) for r in cur.fetchall()])

    signals.sort(key=lambda x: x.get("date") or "0", reverse=True)
    return signals


def _fetch_outreach(conn, days_back=30):
    cur = conn.cursor()
    cur.execute("""
        SELECT o.outreach_date, c.name as company,
               COALESCE(ct.first_name || ' ' || ct.last_name, '-') as contact,
               o.outreach_type, o.outcome, o.notes,
               o.follow_up_date,
               CASE WHEN o.follow_up_done = 1 THEN 'Yes' ELSE 'No' END as followed_up
        FROM outreach_log o
        JOIN companies c ON o.target_company_id = c.id
        LEFT JOIN contacts ct ON o.target_contact_id = ct.id
        WHERE o.outreach_date >= date('now', ? || ' days')
        ORDER BY o.outreach_date DESC
    """, (f"-{days_back}",))
    return [dict(r) for r in cur.fetchall()]


def _fetch_top_opportunities(conn, limit=20):
    cur = conn.cursor()
    cur.execute("""
        SELECT c.name, c.status, c.industry,
               COALESCE(c.opportunity_score, 0) as opp_score,
               COUNT(DISTINCT ct.id) as contact_count,
               MAX(o.outreach_date) as last_touch,
               COUNT(DISTINCT f.id) as funding_events,
               COUNT(DISTINCT h.id) as hiring_signals
        FROM companies c
        LEFT JOIN contacts ct ON c.id = ct.company_id
        LEFT JOIN outreach_log o ON c.id = o.target_company_id
        LEFT JOIN funding_events f ON c.id = f.company_id
        LEFT JOIN hiring_signals h ON c.id = h.company_id
        WHERE c.status IN ('high_growth_target', 'prospect')
        GROUP BY c.id
        ORDER BY COALESCE(c.opportunity_score, 0) DESC
        LIMIT ?
    """, (limit,))
    return [dict(r) for r in cur.fetchall()]


def _fetch_market_notes(conn, days_back=30):
    cur = conn.cursor()
    cur.execute("""
        SELECT note_date, note_text, source,
               COALESCE(companies_mentioned, '') as companies,
               COALESCE(tags, '') as tags
        FROM market_notes
        WHERE note_date >= date('now', ? || ' days')
        ORDER BY note_date DESC
    """, (f"-{days_back}",))
    return [dict(r) for r in cur.fetchall()]


def _fetch_kpis(conn):
    cur = conn.cursor()
    kpis = {}

    cur.execute("SELECT COUNT(*) FROM companies WHERE status IN ('high_growth_target', 'prospect')")
    kpis["active_targets"] = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM contacts")
    kpis["total_contacts"] = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM outreach_log WHERE outreach_date >= date('now', '-7 days')")
    kpis["weekly_touches"] = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM outreach_log WHERE outreach_date >= date('now', '-30 days')")
    kpis["monthly_touches"] = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM funding_events WHERE event_date >= date('now', '-30 days')")
    kpis["funding_30d"] = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM hiring_signals WHERE signal_date >= date('now', '-30 days')")
    kpis["hiring_30d"] = cur.fetchone()[0]

    cur.execute("""
        SELECT COUNT(*) FROM companies
        WHERE status = 'active_client'
        AND NOT EXISTS (
            SELECT 1 FROM outreach_log o
            WHERE o.target_company_id = companies.id
            AND o.outreach_date >= date('now', '-60 days')
        )
    """)
    kpis["at_risk"] = cur.fetchone()[0]

    return kpis


# =============================================================================
# SHEET BUILDERS
# =============================================================================

def _build_overview(wb, kpis):
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_properties.tabColor = "2F5496"

    today = datetime.now()
    ws["A1"] = "Relationship Engine — Weekly Report"
    ws["A1"].font = FONT_TITLE
    ws["A2"] = f"Generated: {today.strftime('%B %d, %Y %I:%M %p')}"
    ws["A2"].font = FONT_DATE

    row = 4
    metrics = [
        ("Active Targets", kpis["active_targets"]),
        ("Total Contacts", kpis["total_contacts"]),
        ("Touches (7d)", kpis["weekly_touches"]),
        ("Touches (30d)", kpis["monthly_touches"]),
        ("Funding Events (30d)", kpis["funding_30d"]),
        ("Hiring Signals (30d)", kpis["hiring_30d"]),
        ("Clients at Risk", kpis["at_risk"]),
    ]

    ws.cell(row=row, column=1, value="Metric").font = FONT_SUBHEADER
    ws.cell(row=row, column=2, value="Value").font = FONT_SUBHEADER
    ws.cell(row=row, column=1).fill = FILL_SECTION
    ws.cell(row=row, column=2).fill = FILL_SECTION

    for i, (label, val) in enumerate(metrics):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).font = FONT_BODY
        ws.cell(row=r, column=2, value=val).font = FONT_BLUE
        ws.cell(row=r, column=2).alignment = ALIGN_CENTER
        if label == "Clients at Risk" and val > 0:
            ws.cell(row=r, column=2).fill = FILL_HIGH

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15


def _build_pipeline(wb, pipeline_data):
    ws = wb.create_sheet("Pipeline")
    ws.sheet_properties.tabColor = "4472C4"

    headers = ["Company", "Status", "Industry", "Opp Score", "Last Outreach",
               "Last Outcome", "Total Touches", "Deal Type", "Deal Status"]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, row_data in enumerate(pipeline_data):
        r = i + 2
        ws.cell(row=r, column=1, value=row_data["name"])
        ws.cell(row=r, column=2, value=(row_data["status"] or "").replace("_", " ").title())
        ws.cell(row=r, column=3, value=row_data.get("industry") or "")
        ws.cell(row=r, column=4, value=row_data["opp_score"])
        ws.cell(row=r, column=4).number_format = "0.0"
        ws.cell(row=r, column=4).alignment = ALIGN_CENTER
        ws.cell(row=r, column=5, value=row_data.get("last_outreach") or "Never")
        ws.cell(row=r, column=6, value=(row_data.get("last_outcome") or "").replace("_", " "))
        ws.cell(row=r, column=7, value=row_data.get("total_touches", 0))
        ws.cell(row=r, column=7).alignment = ALIGN_CENTER
        ws.cell(row=r, column=8, value=row_data.get("deal_type") or "")
        ws.cell(row=r, column=9, value=(row_data.get("deal_status") or "").replace("_", " "))
        _style_data_row(ws, r, len(headers), alt=(i % 2 == 1))

        score = row_data["opp_score"]
        if score >= 30:
            ws.cell(row=r, column=4).fill = FILL_LOW
        elif score >= 10:
            ws.cell(row=r, column=4).fill = FILL_MED

    _auto_width(ws)


def _build_signals(wb, signals):
    ws = wb.create_sheet("Signals")
    ws.sheet_properties.tabColor = "ED7D31"

    headers = ["Date", "Type", "Company", "Detail", "Extra", "Company Status"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, sig in enumerate(signals):
        r = i + 2
        ws.cell(row=r, column=1, value=sig.get("date", ""))
        ws.cell(row=r, column=2, value=sig.get("type", ""))
        ws.cell(row=r, column=3, value=sig.get("name", ""))
        ws.cell(row=r, column=4, value=sig.get("detail", ""))
        ws.cell(row=r, column=5, value=sig.get("extra", ""))
        ws.cell(row=r, column=6, value=(sig.get("status") or "").replace("_", " ").title())
        _style_data_row(ws, r, len(headers), alt=(i % 2 == 1))

        sig_type = sig.get("type", "")
        if sig_type == "Funding":
            ws.cell(row=r, column=2).fill = FILL_LOW
        elif sig_type == "Hiring":
            ws.cell(row=r, column=2).fill = FILL_MED
        elif sig_type == "Lease Expiry":
            ws.cell(row=r, column=2).fill = FILL_HIGH

    _auto_width(ws)


def _build_outreach(wb, outreach):
    ws = wb.create_sheet("Outreach")
    ws.sheet_properties.tabColor = "70AD47"

    headers = ["Date", "Company", "Contact", "Type", "Outcome", "Notes",
               "Follow-Up Date", "Followed Up"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, o in enumerate(outreach):
        r = i + 2
        ws.cell(row=r, column=1, value=o.get("outreach_date", ""))
        ws.cell(row=r, column=2, value=o.get("company", ""))
        ws.cell(row=r, column=3, value=o.get("contact", ""))
        ws.cell(row=r, column=4, value=o.get("outreach_type", ""))
        ws.cell(row=r, column=5, value=(o.get("outcome") or "").replace("_", " "))
        ws.cell(row=r, column=6, value=o.get("notes") or "")
        ws.cell(row=r, column=6).alignment = ALIGN_LEFT
        ws.cell(row=r, column=7, value=o.get("follow_up_date") or "")
        ws.cell(row=r, column=8, value=o.get("followed_up", ""))
        _style_data_row(ws, r, len(headers), alt=(i % 2 == 1))

        outcome = o.get("outcome", "")
        if outcome in ("meeting_booked", "meeting_held", "deal_started"):
            ws.cell(row=r, column=5).fill = FILL_LOW
        elif outcome in ("no_response", "declined"):
            ws.cell(row=r, column=5).fill = FILL_HIGH
        elif outcome == "pending":
            ws.cell(row=r, column=5).fill = FILL_MED

    _auto_width(ws)


def _build_opportunities(wb, opportunities):
    ws = wb.create_sheet("Top Opportunities")
    ws.sheet_properties.tabColor = "FFC000"

    headers = ["Rank", "Company", "Status", "Industry", "Opp Score",
               "Contacts", "Last Touch", "Funding Events", "Hiring Signals"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, opp in enumerate(opportunities):
        r = i + 2
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=1).alignment = ALIGN_CENTER
        ws.cell(row=r, column=2, value=opp["name"])
        ws.cell(row=r, column=3, value=(opp["status"] or "").replace("_", " ").title())
        ws.cell(row=r, column=4, value=opp.get("industry") or "")
        ws.cell(row=r, column=5, value=opp["opp_score"])
        ws.cell(row=r, column=5).number_format = "0.0"
        ws.cell(row=r, column=5).alignment = ALIGN_CENTER
        ws.cell(row=r, column=6, value=opp.get("contact_count", 0))
        ws.cell(row=r, column=6).alignment = ALIGN_CENTER
        ws.cell(row=r, column=7, value=opp.get("last_touch") or "Never")
        ws.cell(row=r, column=8, value=opp.get("funding_events", 0))
        ws.cell(row=r, column=8).alignment = ALIGN_CENTER
        ws.cell(row=r, column=9, value=opp.get("hiring_signals", 0))
        ws.cell(row=r, column=9).alignment = ALIGN_CENTER
        _style_data_row(ws, r, len(headers), alt=(i % 2 == 1))

    _auto_width(ws)


def _build_market_notes(wb, notes):
    ws = wb.create_sheet("Market Notes")
    ws.sheet_properties.tabColor = "7030A0"

    headers = ["Date", "Note", "Source", "Companies Mentioned", "Tags"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, n in enumerate(notes):
        r = i + 2
        ws.cell(row=r, column=1, value=n.get("note_date", ""))
        ws.cell(row=r, column=2, value=n.get("note_text", ""))
        ws.cell(row=r, column=2).alignment = ALIGN_LEFT
        ws.cell(row=r, column=3, value=n.get("source") or "")
        ws.cell(row=r, column=4, value=n.get("companies", "").replace("|", ", "))
        ws.cell(row=r, column=5, value=n.get("tags", "").replace("|", ", "))
        _style_data_row(ws, r, len(headers), alt=(i % 2 == 1))

    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["D"].width = 25
    _auto_width(ws, min_width=12)
    ws.column_dimensions["B"].width = 60


# =============================================================================
# MAIN
# =============================================================================

def generate_report(
    output_path: str = None,
    days_back: int = 30,
    db_path: str = None
) -> str:
    """
    Generate a full weekly Excel report.

    Args:
        output_path: Where to save the .xlsx. Defaults to private_data/reports/
        days_back: How many days of data to include
        db_path: Database path override

    Returns:
        Path to the generated .xlsx file
    """
    if output_path is None:
        report_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "private_data", "reports"
        )
        os.makedirs(report_dir, exist_ok=True)
        today = datetime.now().strftime("%Y-%m-%d")
        output_path = os.path.join(report_dir, f"re_report_{today}.xlsx")

    conn = _get_conn(db_path)

    kpis = _fetch_kpis(conn)
    pipeline = _fetch_pipeline(conn)
    signals = _fetch_signals(conn, days_back)
    outreach = _fetch_outreach(conn, days_back)
    opportunities = _fetch_top_opportunities(conn, 20)
    notes = _fetch_market_notes(conn, days_back)

    conn.close()

    wb = Workbook()
    _build_overview(wb, kpis)
    _build_pipeline(wb, pipeline)
    _build_signals(wb, signals)
    _build_outreach(wb, outreach)
    _build_opportunities(wb, opportunities)
    _build_market_notes(wb, notes)

    wb.save(output_path)
    print(f"Report saved: {output_path}")
    return output_path


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Generate Excel report")
    parser.add_argument("--output", "-o", help="Output file path")
    parser.add_argument("--days", "-d", type=int, default=30, help="Days of data to include")
    args = parser.parse_args()
    generate_report(output_path=args.output, days_back=args.days)
